import os
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

def plot_excel_data(file_path):
    """
    Plot data from an Excel file.

    Parameters:
    - file_path (str): Path to the Excel file or 'AM' to plot area numbers.

    If file_path is 'AM', it plots area numbers. Otherwise, it analyzes the data
    from the Excel file and plots accordingly.
    """

    # Define warm colors
    warm_color1 = (255/255, 202/255, 158/255)  # Light orange
    warm_color2 = (255/255, 174/255, 149/255)  # Light red

    # Create a new figure and axis
    fig, ax = plt.subplots()

    # Plot area numbers if file_path is 'AM'
    if file_path == 'AM':
        # Define x and y axis ticks
        ax.set_xticks(range(6))
        ax.set_yticks(range(6))

        # Define x and y axis labels
        ax.set_xticklabels(['0%', '20%', '40%', '60%', '80%', '100%'])
        ax.set_yticklabels(['0%', '20%', '40%', '60%', '80%', '100%'])

        # Loop through all areas and fill colors
        for i in range(5):
            for j in range(5):
                if (i + j) % 2 == 0:
                    ax.add_patch(plt.Rectangle((i, j), 1, 1, color=warm_color1))
                else:
                    ax.add_patch(plt.Rectangle((i, j), 1, 1, color=warm_color2))
                # The program output column will display the species codes for each region
                tem = j * 5 + i + 1  # (f)[j*5+i+1]
                # tem = j*5+i+1 #Show area codes
                ax.text(i + 0.5, j + 0.5, f"{tem}", ha='center', va='center', fontsize=10)

    else:
        # Load data from the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        num = [0] * 444
        mat = np.zeros((26, 500))
        gap = 20
        GAP = 0.2
        f = np.zeros(26)

        min_row = 2
        max_row = sheet.max_row
        min_col = 1
        max_col = sheet.max_column

        per_num = max_col - min_col + 1

        # Loop through different ranges of data
        for a in range(1, 6):
            for b in range(1, 6):
                k = 0
                # Iterate over rows in the specified range
                for i in sheet.iter_rows(min_row, max_row, min_col, max_col):
                    for j in i:
                        if (float(j.value) <= gap * a and float(j.value) > gap * (a - 1)):
                            num[k] = num[k] + 1
                    if ((num[k] <= per_num * GAP * b and num[k] > per_num * (b - 1) * GAP)):
                        mat[(a - 1) * 5 + b][k] = j.row
                    num[k] = 0
                    k = k + 1

        # Count occurrences of data in each cell
        for a in range(1, 6):
            for b in range(1, 6):
                for k in range(0, 499):
                    if (mat[(a - 1) * 5 + b][k] > 0):
                        f[(a - 1) * 5 + b] = f[(a - 1) * 5 + b] + 1

        # Write mat-related output to a text file
        output_file_path = "mat_output.txt"
        with open(output_file_path, "w") as txt_file:
            txt_file.write("MAT Output:\n")
            for a in range(1, 6):
                for b in range(1, 6):
                    txt_file.write(f"mat[{(a - 1) * 5 + b}]\n")
                    for k in range(0, 499):
                        if (mat[(a - 1) * 5 + b][k] > 0):
                            txt_file.write(f"{mat[(a - 1) * 5 + b][k]} ")
                    txt_file.write("\n")

        # Loop through all areas and fill colors
        for i in range(5):
            for j in range(5):
                if (i + j) % 2 == 0:
                    ax.add_patch(plt.Rectangle((i, j), 1, 1, color=warm_color1))
                else:
                    ax.add_patch(plt.Rectangle((i, j), 1, 1, color=warm_color2))
                # The program output column will display the species codes for each region
                tem = (f)[j*5+i+1]
                # tem = j*5+i+1 #Show area codes
                ax.text(i + 0.5, j + 0.5, f"{tem}", ha='center', va='center', fontsize=10)

    # Set x and y axis labels
    plt.xlabel('Proportion of a certain type of bacteria in the human intestinal flora')
    plt.ylabel('Proportion of patients carrying a certain type of bacteria in the overall population')

    # Set the range of the chart
    ax.set_xlim(0, 5)
    ax.set_ylim(0, 5)

    # Display the chart
    plt.gca().set_aspect('equal', adjustable='box')
    plt.grid(True)
    plt.show()

plot_excel_data('D:\\file\works\MSM_ALL.xlsx')
